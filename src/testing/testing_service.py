import json
import time
import uuid
from datetime import datetime
from typing import List, Optional

from src.domain.services.dataset_service import DatasetService
from src.domain.services.llm_dto import LLMQuestion
from src.domain.services.llm_service import LLMService
from src.infrastructure.logger import get_prefixed_logger
from src.infrastructure.paths import PROJECT_ROOT
from src.testing.testing_dto import (
    TestQuestion,
    TestConfig,
    TestResult,
    TestReport,
    BulkTestRequest,
    DatasetResultItem,
)
from src.vector_search.embedder import embed_batch_with_ids

logger = get_prefixed_logger(__name__, "TESTING_SERVICE")


# =============================================================================
# AUEPORA RETRIEVAL METRICS
# Based on: "Evaluation of Retrieval-Augmented Generation: A Survey" (Yu et al., 2024)
# =============================================================================

def precision_at_k(
    datasets: List[DatasetResultItem],
    k: int,
    threshold: float = 0.5
) -> float:
    """
    Precision@k: fraction of relevant documents among top-k retrieved.

    Precision@k = |relevant ∩ top-k| / k

    Args:
        datasets: Ranked list of retrieved datasets
        k: Number of top results to consider
        threshold: Minimum relevance_rating to be considered relevant (default 0.5)

    Returns:
        Precision score between 0 and 1
    """
    if k <= 0:
        return 0.0
    top_k = datasets[:k]
    relevant_count = sum(
        1 for d in top_k
        if d.relevance_rating is not None and d.relevance_rating >= threshold
    )
    return relevant_count / k


def recall_at_k(
    datasets: List[DatasetResultItem],
    expected_datasets: Optional[dict[str, float]],
    k: int,
    threshold: float = 0.5
) -> float:
    """
    Recall@k: fraction of relevant documents retrieved in top-k.

    Recall@k = |relevant ∩ top-k| / |all relevant|

    Args:
        datasets: Ranked list of retrieved datasets
        expected_datasets: Ground truth dict {dataset_id: relevance_rating}
        k: Number of top results to consider
        threshold: Minimum relevance_rating to be considered relevant

    Returns:
        Recall score between 0 and 1
    """
    if not expected_datasets:
        return 0.0

    # Get all relevant documents from ground truth
    relevant_gt = {
        did for did, rating in expected_datasets.items()
        if rating >= threshold
    }

    if not relevant_gt:
        return 0.0

    # Get retrieved document IDs in top-k
    top_k_ids = {d.dataset_id for d in datasets[:k]}

    # Calculate intersection
    retrieved_relevant = len(top_k_ids & relevant_gt)

    return retrieved_relevant / len(relevant_gt)


def reciprocal_rank(
    datasets: List[DatasetResultItem],
    threshold: float = 0.5
) -> float:
    """
    Reciprocal Rank: 1/position of the first relevant document.

    RR = 1 / rank_first_relevant

    Args:
        datasets: Ranked list of retrieved datasets
        threshold: Minimum relevance_rating to be considered relevant

    Returns:
        Reciprocal rank between 0 and 1 (0 if no relevant found)
    """
    for i, d in enumerate(datasets):
        if d.relevance_rating is not None and d.relevance_rating >= threshold:
            return 1.0 / (i + 1)
    return 0.0


def average_precision_at_k(
    datasets: List[DatasetResultItem],
    k: int,
    threshold: float = 0.5
) -> float:
    """
    Average Precision@k: mean of precision values at each relevant position.

    AP@k = (1/|relevant|) × Σ(P(i) × rel(i))

    Args:
        datasets: Ranked list of retrieved datasets
        k: Number of top results to consider
        threshold: Minimum relevance_rating to be considered relevant

    Returns:
        Average precision score between 0 and 1
    """
    top_k = datasets[:k]
    precisions_at_relevant = []
    relevant_count = 0

    for i, d in enumerate(top_k):
        if d.relevance_rating is not None and d.relevance_rating >= threshold:
            relevant_count += 1
            precision_at_i = relevant_count / (i + 1)
            precisions_at_relevant.append(precision_at_i)

    if not precisions_at_relevant:
        return 0.0

    return sum(precisions_at_relevant) / len(precisions_at_relevant)


def hit_at_k(
    datasets: List[DatasetResultItem],
    k: int,
    threshold: float = 0.5
) -> int:
    """
    Hit@k: binary indicator if any relevant document is in top-k.

    Hit@k = 1 if |relevant ∩ top-k| > 0, else 0

    Args:
        datasets: Ranked list of retrieved datasets
        k: Number of top results to consider
        threshold: Minimum relevance_rating to be considered relevant

    Returns:
        1 if hit, 0 otherwise
    """
    top_k = datasets[:k]
    for d in top_k:
        if d.relevance_rating is not None and d.relevance_rating >= threshold:
            return 1
    return 0


def ndcg_at_k(
    datasets: List[DatasetResultItem],
    k: int
) -> float:
    """
    Normalized Discounted Cumulative Gain@k: measures ranking quality
    considering graded relevance.

    DCG@k = Σ(rel_i / log2(i+1))
    NDCG@k = DCG@k / IDCG@k

    Args:
        datasets: Ranked list of retrieved datasets
        k: Number of top results to consider

    Returns:
        NDCG score between 0 and 1
    """
    import math

    top_k = datasets[:k]

    # Calculate DCG
    dcg = 0.0
    for i, d in enumerate(top_k):
        rel = d.relevance_rating if d.relevance_rating is not None else 0.0
        # Using formula: rel_i / log2(i + 2) to avoid log(1) = 0
        dcg += rel / math.log2(i + 2)

    # Calculate IDCG (ideal DCG with perfect ranking)
    ideal_rels = sorted(
        [d.relevance_rating if d.relevance_rating is not None else 0.0 for d in top_k],
        reverse=True
    )
    idcg = 0.0
    for i, rel in enumerate(ideal_rels):
        idcg += rel / math.log2(i + 2)

    if idcg == 0:
        return 0.0

    return dcg / idcg

# File paths
TEST_DATA_DIR = PROJECT_ROOT / "test_data"
QUESTIONS_FILE = TEST_DATA_DIR / "questions.json"
REPORTS_DIR = TEST_DATA_DIR / "reports"


class TestingService:
    """Service for bulk testing of dataset search"""

    def __init__(self, dataset_service: DatasetService, llm_service: LLMService):
        self.dataset_service = dataset_service
        self.llm_service = llm_service
        self._ensure_directories()

    def _ensure_directories(self):
        """Ensure test data directories exist"""
        TEST_DATA_DIR.mkdir(exist_ok=True)
        REPORTS_DIR.mkdir(exist_ok=True)

        # Create empty questions file if it doesn't exist
        if not QUESTIONS_FILE.exists():
            QUESTIONS_FILE.write_text("[]")
            logger.info(f"Created empty questions file: {QUESTIONS_FILE}")

    # region Questions Management

    def add_question(
        self,
        question: str,
        city: Optional[str] = None,
        state: Optional[str] = None,
        country: Optional[str] = None,
        year_from: Optional[int] = None,
        year_to: Optional[int] = None,
        expected_datasets: Optional[dict[str, float]] = None,
    ) -> TestQuestion:
        """Add a new test question"""
        questions = self.get_all_questions()

        new_question = TestQuestion(
            id=str(uuid.uuid4()),
            question=question,
            city=city,
            state=state,
            country=country,
            year_from=year_from,
            year_to=year_to,
            expected_datasets=expected_datasets,
        )

        questions.append(new_question)
        self._save_questions(questions)

        logger.info(
            f"Added new test question: {question} (city={city}, state={state}, country={country}, "
            f"year_from={year_from}, year_to={year_to}, expected_datasets={expected_datasets})"
        )
        return new_question

    def get_all_questions(self) -> List[TestQuestion]:
        """Get all test questions"""
        try:
            with open(QUESTIONS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return [TestQuestion(**q) for q in data]
        except Exception as e:
            logger.error(f"Error loading questions: {e}")
            return []

    def _save_questions(self, questions: List[TestQuestion]):
        """Save questions to file"""
        with open(QUESTIONS_FILE, "w", encoding="utf-8") as f:
            json.dump(
                [q.model_dump(mode="json") for q in questions],
                f,
                ensure_ascii=False,
                indent=2,
            )

    # endregion

    # region Test Execution

    async def run_single_test(
        self,
        question: str,
        config: TestConfig,
        use_multi_query: bool,
        city_filter: Optional[str] = None,
        state_filter: Optional[str] = None,
        country_filter: Optional[str] = None,
        year_from: Optional[int] = None,
        year_to: Optional[int] = None,
        expected_datasets: Optional[dict[str, float]] = None,
    ) -> TestResult:
        """Execute a single test with given configuration"""
        start_time = time.perf_counter()
        research_questions = None

        try:
            # Step 1: Generate research questions if multi-query enabled
            if use_multi_query:
                research_questions_objs = await self.llm_service.get_research_questions(
                    question
                )
                research_questions = [q.question for q in research_questions_objs]
            else:
                research_questions_objs = [
                    LLMQuestion(question=question, reason="Direct user query")
                ]

            # Step 2: Generate embeddings with specified embedder model
            questions_list = [
                {"text": q.question, "id": q.question_hash}
                for q in research_questions_objs
            ]
            embeddings = await embed_batch_with_ids(
                questions_list, embedder_model=config.embedder_model
            )

            # Step 3: Search for each research question using specified embedder
            # Note: search uses maximum accuracy settings (ef=256) automatically
            all_datasets = []
            for emb in embeddings:
                datasets = await self.dataset_service.search_datasets_with_embeddings(
                    emb["embedding"],
                    embedder_model=config.embedder_model,
                    city_filter=city_filter,
                    state_filter=state_filter,
                    country_filter=country_filter,
                    year_from=year_from,
                    year_to=year_to,
                )
                all_datasets.extend(datasets[: config.limit])

            # Remove duplicates by dataset ID
            seen_ids = set()
            unique_datasets = []
            for ds in all_datasets:
                ds_id = ds.metadata.id
                if ds_id not in seen_ids:
                    seen_ids.add(ds_id)
                    unique_datasets.append(ds)

            # Limit to config.limit
            datasets_found = unique_datasets[: config.limit]

            execution_time = time.perf_counter() - start_time

            return TestResult(
                question=question,
                config=config,
                datasets_found=len(datasets_found),
                datasets=[
                    DatasetResultItem(
                        title=ds.metadata.title,
                        score=ds.score,
                        dataset_id=ds.metadata.id,
                        relevance_rating=(
                            expected_datasets.get(ds.metadata.id)
                            if expected_datasets
                            else None
                        ),
                    )
                    for ds in datasets_found
                ],
                execution_time_seconds=execution_time,
                research_questions=research_questions,
                error=None,
                applied_city_filter=city_filter,
                applied_state_filter=state_filter,
                applied_country_filter=country_filter,
                applied_year_from=year_from,
                applied_year_to=year_to,
                used_multi_query=use_multi_query,
            )

        except Exception as e:
            execution_time = time.perf_counter() - start_time
            logger.error(f"Test failed for question '{question}': {e}", exc_info=True)
            return TestResult(
                question=question,
                config=config,
                datasets_found=0,
                datasets=[],
                execution_time_seconds=execution_time,
                research_questions=research_questions,
                error=str(e),
                applied_city_filter=city_filter,
                applied_state_filter=state_filter,
                applied_country_filter=country_filter,
                applied_year_from=year_from,
                applied_year_to=year_to,
                used_multi_query=use_multi_query,
            )

    async def run_bulk_test(self, request: BulkTestRequest) -> TestReport:
        """Execute bulk testing with multiple configurations"""
        logger.info("Starting bulk test execution")
        start_time = time.perf_counter()

        # Get questions to test
        all_questions = self.get_all_questions()
        if request.question_indices:
            questions_to_test = [
                all_questions[i]
                for i in request.question_indices
                if i < len(all_questions)
            ]
        else:
            questions_to_test = all_questions

        # Determine which variants to run based on filters and multiquery parameters
        variants_to_run = []

        # Variant 1: WITH filters + WITH multiquery
        if (request.filters is None or request.filters is True) and \
           (request.multiquery is None or request.multiquery is True):
            variants_to_run.append((True, True))

        # Variant 2: WITH filters + WITHOUT multiquery
        if (request.filters is None or request.filters is True) and \
           (request.multiquery is None or request.multiquery is False):
            variants_to_run.append((True, False))

        # Variant 3: WITHOUT filters + WITH multiquery
        if (request.filters is None or request.filters is False) and \
           (request.multiquery is None or request.multiquery is True):
            variants_to_run.append((False, True))

        # Variant 4: WITHOUT filters + WITHOUT multiquery
        if (request.filters is None or request.filters is False) and \
           (request.multiquery is None or request.multiquery is False):
            variants_to_run.append((False, False))

        enabled_variants = len(variants_to_run)
        total_tests = len(questions_to_test) * len(request.test_configs) * enabled_variants

        logger.info(
            f"Testing {len(questions_to_test)} questions with {len(request.test_configs)} configurations "
            f"in {enabled_variants} variants each = {total_tests} total tests"
        )

        # Run all tests
        results = []
        current_test = 0

        for question in questions_to_test:
            for config in request.test_configs:
                for use_filters, use_multiquery in variants_to_run:
                    current_test += 1

                    # Prepare variant description
                    filters_desc = "WITH filters" if use_filters else "WITHOUT filters"
                    multiquery_desc = "WITH multi-query" if use_multiquery else "WITHOUT multi-query"

                    logger.info(
                        f"Running test {current_test}/{total_tests} ({filters_desc} + {multiquery_desc})"
                    )

                    result = await self.run_single_test(
                        question.question,
                        config,
                        use_multi_query=use_multiquery,
                        city_filter=question.city if use_filters else None,
                        state_filter=question.state if use_filters else None,
                        country_filter=question.country if use_filters else None,
                        year_from=question.year_from if use_filters else None,
                        year_to=question.year_to if use_filters else None,
                        expected_datasets=question.expected_datasets,
                    )
                    results.append(result)

        # Calculate statistics
        execution_time = time.perf_counter() - start_time
        successful_tests = sum(1 for r in results if r.error is None)
        failed_tests = sum(1 for r in results if r.error is not None)

        # Create report
        report = TestReport(
            report_id=str(uuid.uuid4()),
            created_at=datetime.now(),
            total_tests=total_tests,
            successful_tests=successful_tests,
            failed_tests=failed_tests,
            total_execution_time_seconds=execution_time,
            results=results,
        )

        # Save report
        self._save_report(report)

        logger.info(
            f"Bulk test completed: {successful_tests}/{total_tests} successful in {execution_time:.2f}s"
        )
        return report

    # endregion

    # region Reports Management

    def _save_report(self, report: TestReport):
        """Save test report to file with timestamp-based filename"""
        # Format: 2025_12_26_01_28.json
        timestamp = report.created_at.strftime("%Y_%m_%d_%H_%M")
        report_file = REPORTS_DIR / f"{timestamp}.json"

        # If file exists, add seconds to make it unique
        if report_file.exists():
            timestamp = report.created_at.strftime("%Y_%m_%d_%H_%M_%S")
            report_file = REPORTS_DIR / f"{timestamp}.json"

        with open(report_file, "w", encoding="utf-8") as f:
            json.dump(report.model_dump(mode="json"), f, ensure_ascii=False, indent=2)
        logger.info(f"Saved report: {report_file}")

    def get_all_reports(self) -> List[dict]:
        """Get list of all reports (metadata only)"""
        reports = []
        for report_file in REPORTS_DIR.glob("*.json"):
            try:
                with open(report_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    reports.append(
                        {
                            "report_id": data["report_id"],
                            "created_at": data["created_at"],
                            "total_tests": data["total_tests"],
                            "successful_tests": data["successful_tests"],
                            "failed_tests": data["failed_tests"],
                        }
                    )
            except Exception as e:
                logger.error(f"Error loading report {report_file}: {e}")

        # Sort by creation time (newest first)
        reports.sort(key=lambda x: x["created_at"], reverse=True)
        return reports

    def get_report(self, report_id: str) -> Optional[TestReport]:
        """Get specific test report by searching through all reports"""
        # Search through all report files to find the one with matching report_id
        for report_file in REPORTS_DIR.glob("*.json"):
            try:
                with open(report_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if data.get("report_id") == report_id:
                        return TestReport(**data)
            except Exception as e:
                logger.error(f"Error loading report {report_file}: {e}")
                continue

        logger.warning(f"Report {report_id} not found")
        return None

    # endregion

    # region Relevance Rating

    def update_relevance_rating(
        self, report_id: str, question: str, dataset_id: str, relevance_rating: float
    ) -> bool:
        """Update relevance rating for a specific dataset in a report"""
        report = self.get_report(report_id)
        if report is None:
            logger.error(f"Report {report_id} not found")
            return False

        # Find the test result for this question
        updated = False
        for result in report.results:
            if result.question == question:
                # Find the dataset in the results
                for dataset in result.datasets:
                    if dataset.dataset_id == dataset_id:
                        dataset.relevance_rating = relevance_rating
                        updated = True
                        break

        if updated:
            # Save the updated report
            self._save_report(report)
            logger.info(
                f"Updated relevance rating for dataset {dataset_id} in question '{question}' to {relevance_rating}"
            )
            return True
        else:
            logger.warning(
                f"Dataset {dataset_id} not found in question '{question}' for report {report_id}"
            )
            return False

    # endregion

    # region Excel Export

    def export_to_excel(self, report_ids: list[str], output_file: str = None) -> str:
        """
        Export multiple reports to Excel for comparison.

        Creates five sheets:
        1. Weighted Scores: avg(score * relevance_rating) per question per experiment
        2. Relevance Metrics: % of relevant datasets per question per experiment
        3. Normalized Weighted: avg(normalized_score * relevance_rating) per question per experiment
           - Scores normalized to 0-100 using global min/max across ALL questions in the experiment
           - Then multiplied by relevance_rating
           - Use this to compare performance across questions within same experiment
        4. Normalized Per Question: avg(normalized_score * relevance_rating) per question per experiment
           - Scores normalized to 0-100 using min/max within EACH question separately
           - Then multiplied by relevance_rating
           - Use this to compare experiments within same question
        5. Detailed Ratings: all datasets with their scores and relevance ratings

        Args:
            report_ids: List of report IDs to compare
            output_file: Optional output file path (default: auto-generated in test_data/reports/)

        Returns:
            Path to the generated Excel file
        """
        try:
            import pandas as pd
            from pathlib import Path
        except ImportError:
            raise ImportError(
                "pandas is required for Excel export. Install with: pip install pandas openpyxl"
            )

        # Load all reports
        reports = []
        for report_id in report_ids:
            report = self.get_report(report_id)
            if report:
                reports.append(report)
            else:
                logger.warning(f"Report {report_id} not found, skipping")

        if not reports:
            raise ValueError("No valid reports found")

        # Collect all unique questions
        all_questions = set()
        for report in reports:
            for result in report.results:
                all_questions.add(result.question)

        all_questions = sorted(list(all_questions))

        # Collect all unique configurations from all reports
        # Each unique (report_id, config) combination is a separate experiment
        experiments = []  # List of (exp_name, report, config_dict)

        for report in reports:
            # Group results by unique config within this report
            configs_seen = {}
            for result in report.results:
                config = result.config
                # Use applied filters to distinguish with/without filter variants
                # We use a flag instead of specific filter values to group all "with filters" results together
                has_location_filters = bool(
                    result.applied_city_filter
                    or result.applied_state_filter
                    or result.applied_country_filter
                )
                config_key = (
                    config.embedder_model.value,
                    config.limit,
                    result.used_multi_query,  # Use actual multi-query flag from result
                    has_location_filters,  # Flag instead of specific values
                )

                if config_key not in configs_seen:
                    configs_seen[config_key] = (config, result)

                    # Create experiment name
                    exp_name = f"{config.embedder_model.value}_limit{config.limit}"

                    # Add multi-query indicator
                    if result.used_multi_query:
                        exp_name += "_multiquery"
                    else:
                        exp_name += "_singlequery"

                    # Add filter indicator
                    if has_location_filters:
                        exp_name += "_with_filters"
                    else:
                        exp_name += "_no_filters"

                    # Store as dict for easy comparison
                    config_dict = {
                        "embedder_model": config.embedder_model.value,
                        "limit": config.limit,
                        "used_multi_query": result.used_multi_query,
                        "has_location_filters": has_location_filters,
                    }

                    experiments.append((exp_name, report, config_dict))

        # Prepare data for weighted scores sheet
        weighted_scores_data = []
        for question in all_questions:
            row = {"Question": question}

            for exp_name, report, config_dict in experiments:
                # Find result for this question with matching config
                question_result = None
                for result in report.results:
                    if result.question == question:
                        # Check if config matches (including applied filters and multi-query)
                        result_config = result.config
                        result_has_filters = bool(
                            result.applied_city_filter
                            or result.applied_state_filter
                            or result.applied_country_filter
                        )
                        if (
                            result_config.embedder_model.value
                            == config_dict["embedder_model"]
                            and result_config.limit == config_dict["limit"]
                            and result.used_multi_query == config_dict["used_multi_query"]
                            and result_has_filters == config_dict["has_location_filters"]
                        ):
                            question_result = result
                            break

                if question_result and question_result.datasets:
                    # Calculate weighted score
                    weighted_sum = 0
                    count = 0
                    for dataset in question_result.datasets:
                        # Use 1.0 as default if relevance_rating is None
                        rating = (
                            dataset.relevance_rating
                            if dataset.relevance_rating is not None
                            else 1.0
                        )
                        weighted_sum += dataset.score * rating
                        count += 1

                    avg_weighted_score = weighted_sum / count if count > 0 else 0
                    row[exp_name] = round(avg_weighted_score, 4)
                else:
                    row[exp_name] = None

            weighted_scores_data.append(row)

        # Prepare data for relevance metrics sheet
        relevance_metrics_data = []
        for question in all_questions:
            row = {"Question": question}

            for exp_name, report, config_dict in experiments:
                # Find result for this question with matching config
                question_result = None
                for result in report.results:
                    if result.question == question:
                        # Check if config matches (including applied filters and multi-query)
                        result_config = result.config
                        result_has_filters = bool(
                            result.applied_city_filter
                            or result.applied_state_filter
                            or result.applied_country_filter
                        )
                        if (
                            result_config.embedder_model.value
                            == config_dict["embedder_model"]
                            and result_config.limit == config_dict["limit"]
                            and result.used_multi_query == config_dict["used_multi_query"]
                            and result_has_filters == config_dict["has_location_filters"]
                        ):
                            question_result = result
                            break

                if question_result and question_result.datasets:
                    # Calculate average relevance rating as percentage
                    total = len(question_result.datasets)
                    relevance_sum = sum(
                        d.relevance_rating if d.relevance_rating is not None else 1.0
                        for d in question_result.datasets
                    )
                    avg_relevance_pct = (relevance_sum / total * 100) if total > 0 else 0
                    row[exp_name] = round(avg_relevance_pct, 1)
                else:
                    row[exp_name] = None

            relevance_metrics_data.append(row)

        # First, find global min/max scores for each experiment across ALL questions
        experiment_min_max = {}  # exp_name -> (min_score, max_score)

        for exp_name, report, config_dict in experiments:
            all_scores_for_exp = []

            # Collect all scores for this experiment from all questions
            for question in all_questions:
                for result in report.results:
                    if result.question == question:
                        # Check if config matches
                        result_config = result.config
                        result_has_filters = bool(
                            result.applied_city_filter
                            or result.applied_state_filter
                            or result.applied_country_filter
                        )
                        if (
                            result_config.embedder_model.value == config_dict["embedder_model"]
                            and result_config.limit == config_dict["limit"]
                            and result.used_multi_query == config_dict["used_multi_query"]
                            and result_has_filters == config_dict["has_location_filters"]
                        ):
                            # Collect all scores from this result
                            all_scores_for_exp.extend([d.score for d in result.datasets])
                            break

            # Calculate global min/max for this experiment
            if all_scores_for_exp:
                experiment_min_max[exp_name] = (min(all_scores_for_exp), max(all_scores_for_exp))
            else:
                experiment_min_max[exp_name] = (0, 1)  # Default fallback

        # Prepare data for normalized weighted scores sheet
        normalized_weighted_data = []
        for question in all_questions:
            row = {"Question": question}

            for exp_name, report, config_dict in experiments:
                # Find result for this question with matching config
                question_result = None
                for result in report.results:
                    if result.question == question:
                        # Check if config matches (including applied filters and multi-query)
                        result_config = result.config
                        result_has_filters = bool(
                            result.applied_city_filter
                            or result.applied_state_filter
                            or result.applied_country_filter
                        )
                        if (
                            result_config.embedder_model.value
                            == config_dict["embedder_model"]
                            and result_config.limit == config_dict["limit"]
                            and result.used_multi_query == config_dict["used_multi_query"]
                            and result_has_filters == config_dict["has_location_filters"]
                        ):
                            question_result = result
                            break

                if question_result and question_result.datasets:
                    # Get global min/max for this experiment
                    min_score, max_score = experiment_min_max[exp_name]

                    # Calculate normalized weighted score using global min/max
                    normalized_weighted_sum = 0
                    count = 0
                    for dataset in question_result.datasets:
                        # Normalize score to 0-100 range using global min/max
                        if max_score > min_score:
                            normalized_score = ((dataset.score - min_score) / (max_score - min_score)) * 100
                        else:
                            normalized_score = 100  # All scores are the same

                        # Use 1.0 as default if relevance_rating is None
                        rating = (
                            dataset.relevance_rating
                            if dataset.relevance_rating is not None
                            else 1.0
                        )

                        # Multiply normalized score by relevance rating
                        normalized_weighted_sum += normalized_score * rating
                        count += 1

                    avg_normalized_weighted = normalized_weighted_sum / count if count > 0 else 0
                    row[exp_name] = round(avg_normalized_weighted, 2)
                else:
                    row[exp_name] = None

            normalized_weighted_data.append(row)

        # Prepare data for normalized per question scores sheet
        normalized_per_question_data = []
        for question in all_questions:
            row = {"Question": question}

            for exp_name, report, config_dict in experiments:
                # Find result for this question with matching config
                question_result = None
                for result in report.results:
                    if result.question == question:
                        # Check if config matches (including applied filters and multi-query)
                        result_config = result.config
                        result_has_filters = bool(
                            result.applied_city_filter
                            or result.applied_state_filter
                            or result.applied_country_filter
                        )
                        if (
                            result_config.embedder_model.value
                            == config_dict["embedder_model"]
                            and result_config.limit == config_dict["limit"]
                            and result.used_multi_query == config_dict["used_multi_query"]
                            and result_has_filters == config_dict["has_location_filters"]
                        ):
                            question_result = result
                            break

                if question_result and question_result.datasets:
                    # Find min and max scores for THIS QUESTION only
                    scores = [d.score for d in question_result.datasets]
                    min_score = min(scores)
                    max_score = max(scores)

                    # Calculate normalized weighted score using per-question min/max
                    normalized_weighted_sum = 0
                    count = 0
                    for dataset in question_result.datasets:
                        # Normalize score to 0-100 range using per-question min/max
                        if max_score > min_score:
                            normalized_score = ((dataset.score - min_score) / (max_score - min_score)) * 100
                        else:
                            normalized_score = 100  # All scores are the same

                        # Use 1.0 as default if relevance_rating is None
                        rating = (
                            dataset.relevance_rating
                            if dataset.relevance_rating is not None
                            else 1.0
                        )

                        # Multiply normalized score by relevance rating
                        normalized_weighted_sum += normalized_score * rating
                        count += 1

                    avg_normalized_weighted = normalized_weighted_sum / count if count > 0 else 0
                    row[exp_name] = round(avg_normalized_weighted, 2)
                else:
                    row[exp_name] = None

            normalized_per_question_data.append(row)

        # Prepare data for detailed relevance ratings sheet
        detailed_ratings_data = []
        for question in all_questions:
            for exp_name, report, config_dict in experiments:
                # Find result for this question with matching config
                question_result = None
                for result in report.results:
                    if result.question == question:
                        # Check if config matches (including applied filters and multi-query)
                        result_config = result.config
                        result_has_filters = bool(
                            result.applied_city_filter
                            or result.applied_state_filter
                            or result.applied_country_filter
                        )
                        if (
                            result_config.embedder_model.value
                            == config_dict["embedder_model"]
                            and result_config.limit == config_dict["limit"]
                            and result.used_multi_query == config_dict["used_multi_query"]
                            and result_has_filters == config_dict["has_location_filters"]
                        ):
                            question_result = result
                            break

                if question_result and question_result.datasets:
                    for dataset in question_result.datasets:
                        detailed_ratings_data.append(
                            {
                                "Question": question,
                                "Experiment": exp_name,
                                "Dataset": dataset.title,
                                "Dataset ID": dataset.dataset_id,
                                "Score": round(dataset.score, 4),
                                "Relevance Rating": dataset.relevance_rating
                                if dataset.relevance_rating is not None
                                else "Not rated",
                            }
                        )

        # =============================================================================
        # AUEPORA RETRIEVAL METRICS
        # Based on: "Evaluation of Retrieval-Augmented Generation: A Survey" (Yu et al., 2024)
        # =============================================================================

        # Build question -> expected_datasets mapping from test questions
        questions_data = self.get_all_questions()
        question_to_expected = {q.question: q.expected_datasets for q in questions_data}

        # Prepare data for AUEPORA metrics sheet (per question, per experiment)
        auepora_metrics_data = []
        for question in all_questions:
            expected = question_to_expected.get(question, {})

            for exp_name, report, config_dict in experiments:
                # Find result for this question with matching config
                question_result = None
                for result in report.results:
                    if result.question == question:
                        result_config = result.config
                        result_has_filters = bool(
                            result.applied_city_filter
                            or result.applied_state_filter
                            or result.applied_country_filter
                        )
                        if (
                            result_config.embedder_model.value == config_dict["embedder_model"]
                            and result_config.limit == config_dict["limit"]
                            and result.used_multi_query == config_dict["used_multi_query"]
                            and result_has_filters == config_dict["has_location_filters"]
                        ):
                            question_result = result
                            break

                if question_result and question_result.datasets:
                    datasets = question_result.datasets
                    k_values = [5, 10, 25]

                    row = {
                        "Question": question[:80] + "..." if len(question) > 80 else question,
                        "Experiment": exp_name,
                        "Retrieved": len(datasets),
                        "Expected": len(expected) if expected else 0,
                    }

                    # Calculate metrics for each k
                    for k in k_values:
                        row[f"P@{k}"] = round(precision_at_k(datasets, k), 4)
                        row[f"R@{k}"] = round(recall_at_k(datasets, expected, k), 4)
                        row[f"Hit@{k}"] = hit_at_k(datasets, k)
                        row[f"AP@{k}"] = round(average_precision_at_k(datasets, k), 4)
                        row[f"NDCG@{k}"] = round(ndcg_at_k(datasets, k), 4)

                    # MRR (doesn't depend on k)
                    row["MRR"] = round(reciprocal_rank(datasets), 4)

                    auepora_metrics_data.append(row)

        # Calculate mean metrics per experiment (summary)
        auepora_summary_data = []
        for exp_name, report, config_dict in experiments:
            # Filter metrics for this experiment
            exp_metrics = [m for m in auepora_metrics_data if m["Experiment"] == exp_name]

            if exp_metrics:
                summary_row = {
                    "Experiment": exp_name,
                    "Questions": len(exp_metrics),
                }

                # Calculate mean for each metric
                metric_columns = [c for c in exp_metrics[0].keys()
                                  if c not in ["Question", "Experiment", "Retrieved", "Expected"]]

                for col in metric_columns:
                    values = [m[col] for m in exp_metrics if m[col] is not None]
                    if values:
                        mean_val = sum(values) / len(values)
                        summary_row[f"Mean {col}"] = round(mean_val, 4)

                auepora_summary_data.append(summary_row)

        # Create DataFrames
        df_weighted_scores = pd.DataFrame(weighted_scores_data)
        df_relevance_metrics = pd.DataFrame(relevance_metrics_data)
        df_normalized_weighted = pd.DataFrame(normalized_weighted_data)
        df_normalized_per_question = pd.DataFrame(normalized_per_question_data)
        df_detailed_ratings = pd.DataFrame(detailed_ratings_data)
        df_auepora_metrics = pd.DataFrame(auepora_metrics_data)
        df_auepora_summary = pd.DataFrame(auepora_summary_data)

        # Generate output filename if not provided
        if output_file is None:
            from datetime import datetime

            # Use date_id format: YYYY_MM_DD_HH_MM_comparison.xlsx
            timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
            # Add shortened report IDs for reference
            if len(report_ids) == 1:
                output_file = REPORTS_DIR / f"{timestamp}_{report_ids[0][:8]}.xlsx"
            else:
                ids_part = "_".join(rid[:6] for rid in report_ids[:2])
                output_file = REPORTS_DIR / f"{timestamp}_{ids_part}_comparison.xlsx"
        else:
            output_file = Path(output_file)

        # Write to Excel with formatting
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df_weighted_scores.to_excel(
                writer, sheet_name="Weighted Scores", index=False
            )
            df_relevance_metrics.to_excel(
                writer, sheet_name="Relevance Metrics", index=False
            )
            df_normalized_weighted.to_excel(
                writer, sheet_name="Normalized Weighted", index=False
            )
            df_normalized_per_question.to_excel(
                writer, sheet_name="Normalized Per Question", index=False
            )
            df_detailed_ratings.to_excel(
                writer, sheet_name="Detailed Ratings", index=False
            )
            df_auepora_metrics.to_excel(
                writer, sheet_name="AUEPORA Metrics", index=False
            )
            df_auepora_summary.to_excel(
                writer, sheet_name="AUEPORA Summary", index=False
            )

            # Apply color formatting
            workbook = writer.book

            # Format Weighted Scores sheet (0 to 1 scale)
            self._apply_color_scale(
                workbook["Weighted Scores"],
                df_weighted_scores,
                min_value=0,
                max_value=1,
            )

            # Format Relevance Metrics sheet (0% to 100% scale)
            self._apply_color_scale(
                workbook["Relevance Metrics"],
                df_relevance_metrics,
                min_value=0,
                max_value=100,
            )

            # Format Normalized Weighted sheet (0 to 100 scale)
            self._apply_color_scale(
                workbook["Normalized Weighted"],
                df_normalized_weighted,
                min_value=0,
                max_value=100,
            )

            # Format Normalized Per Question sheet (0 to 100 scale)
            self._apply_color_scale(
                workbook["Normalized Per Question"],
                df_normalized_per_question,
                min_value=0,
                max_value=100,
            )

            # Format Score column in Detailed Ratings (0 to 1 scale)
            self._apply_score_colors(workbook["Detailed Ratings"], df_detailed_ratings)

            # Format AUEPORA Metrics sheet (0 to 1 scale for most metrics)
            self._apply_auepora_colors(workbook["AUEPORA Metrics"], df_auepora_metrics)

            # Format AUEPORA Summary sheet (0 to 1 scale for most metrics)
            self._apply_auepora_summary_colors(workbook["AUEPORA Summary"], df_auepora_summary)

        logger.info(f"Excel report exported to: {output_file}")
        return str(output_file)

    def _apply_color_scale(self, worksheet, dataframe, min_value=0, max_value=1):
        """Apply red-to-green color scale to numeric columns (direct cell coloring for Mac Numbers compatibility)"""
        from openpyxl.styles import PatternFill

        # Get the range of data columns (skip Question column)
        first_data_col = 2  # Column B (after Question)
        last_data_col = len(dataframe.columns)
        first_data_row = 2  # Row 2 (after header)
        last_data_row = len(dataframe) + 1  # +1 for header

        if last_data_row < first_data_row:
            return

        # Apply color directly to each cell based on its value
        for row_idx in range(first_data_row, last_data_row + 1):
            for col_idx in range(first_data_col, last_data_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = cell.value

                # Skip non-numeric values
                if value is None or not isinstance(value, (int, float)):
                    continue

                # Calculate color based on value
                color = self._get_color_for_value(value, min_value, max_value)
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

    def _get_color_for_value(self, value, min_value, max_value):
        """Get RGB color for a value in range [min_value, max_value]"""
        # Normalize value to 0-1 range
        if max_value == min_value:
            normalized = 0.5
        else:
            normalized = (value - min_value) / (max_value - min_value)
            normalized = max(0, min(1, normalized))  # Clamp to 0-1

        # Colors: Red (F8696B) -> Yellow (FFEB84) -> Green (63BE7B)
        if normalized <= 0.5:
            # Red to Yellow
            ratio = normalized * 2  # 0 to 1
            r = int(0xF8 + (0xFF - 0xF8) * ratio)
            g = int(0x69 + (0xEB - 0x69) * ratio)
            b = int(0x6B + (0x84 - 0x6B) * ratio)
        else:
            # Yellow to Green
            ratio = (normalized - 0.5) * 2  # 0 to 1
            r = int(0xFF + (0x63 - 0xFF) * ratio)
            g = int(0xEB + (0xBE - 0xEB) * ratio)
            b = int(0x84 + (0x7B - 0x84) * ratio)

        return f"{r:02X}{g:02X}{b:02X}"

    def _apply_score_colors(self, worksheet, dataframe):
        """Apply color scale to Score column in Detailed Ratings (direct cell coloring for Mac Numbers compatibility)"""
        from openpyxl.styles import PatternFill

        # Find Score column index
        try:
            score_col_idx = list(dataframe.columns).index("Score") + 1
        except ValueError:
            return

        first_row = 2
        last_row = len(dataframe) + 1

        if last_row < first_row:
            return

        # Apply color directly to each cell in Score column
        for row_idx in range(first_row, last_row + 1):
            cell = worksheet.cell(row=row_idx, column=score_col_idx)
            value = cell.value

            # Skip non-numeric values
            if value is None or not isinstance(value, (int, float)):
                continue

            # Calculate color based on value (0 to 1 scale)
            color = self._get_color_for_value(value, 0, 1)
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

    def _apply_auepora_colors(self, worksheet, dataframe):
        """Apply color scale to AUEPORA metrics columns (0 to 1 scale)"""
        from openpyxl.styles import PatternFill

        # Columns to color (all metric columns, not Question/Experiment/Retrieved/Expected)
        metric_columns = [
            col for col in dataframe.columns
            if col not in ["Question", "Experiment", "Retrieved", "Expected"]
        ]

        first_row = 2
        last_row = len(dataframe) + 1

        if last_row < first_row:
            return

        for col_name in metric_columns:
            try:
                col_idx = list(dataframe.columns).index(col_name) + 1
            except ValueError:
                continue

            # Determine scale: Hit@k is 0-1 binary, others are 0-1 continuous
            max_value = 1.0

            for row_idx in range(first_row, last_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = cell.value

                if value is None or not isinstance(value, (int, float)):
                    continue

                color = self._get_color_for_value(value, 0, max_value)
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

    def _apply_auepora_summary_colors(self, worksheet, dataframe):
        """Apply color scale to AUEPORA summary columns (0 to 1 scale)"""
        from openpyxl.styles import PatternFill

        # All columns except Experiment and Questions
        metric_columns = [
            col for col in dataframe.columns
            if col not in ["Experiment", "Questions"]
        ]

        first_row = 2
        last_row = len(dataframe) + 1

        if last_row < first_row:
            return

        for col_name in metric_columns:
            try:
                col_idx = list(dataframe.columns).index(col_name) + 1
            except ValueError:
                continue

            for row_idx in range(first_row, last_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = cell.value

                if value is None or not isinstance(value, (int, float)):
                    continue

                color = self._get_color_for_value(value, 0, 1)
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

    # endregion


# region Dependency injection
_testing_service: Optional[TestingService] = None


def get_testing_service(
    dataset_service: DatasetService, llm_service: LLMService
) -> TestingService:
    """Get TestingService instance"""
    global _testing_service
    if _testing_service is None:
        _testing_service = TestingService(dataset_service, llm_service)
    return _testing_service


# endregion
